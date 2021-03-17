# -*- coding: UTF-8 -*-
'''=================================================
@Project -> File   ：python_openpyxl -> 04-pyExcel_row-column
@IDE    ：PyCharm
@Author ：zhangzhen
@Date   ：2020/5/15 9:51
@Desc   ：获取最大行数,列数;展示每行每列的数据
=================================================='''
from openpyxl import load_workbook
wb = load_workbook('./ATtest.xlsx')
ws2 = wb['AT_SMS_SendOrder']  # 常用
#获得最大行
print(ws2.max_row)
#获得最大列
print(ws2.max_column)

'''
获取每一行，每一列
sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹。
sheet.columns类似，不过里面是每个tuple是每一列的单元格。
'''
# 按行，所以返回A1, B1, C1这样的顺序
for row in ws2.rows:
    print(row)
    for cell in row:
        print(cell.value)

# A1, A2, A3这样的顺序
for column in ws2.columns:
    for cell in column:
        print(cell.value)



