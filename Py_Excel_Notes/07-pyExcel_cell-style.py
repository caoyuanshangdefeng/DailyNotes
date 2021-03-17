# -*- coding: UTF-8 -*-
'''=================================================
@Project -> File   ：python_openpyxl -> 07-pyExcel_cell-style
@IDE    ：PyCharm
@Author ：zhangzhen
@Date   ：2020/5/15 10:55
@Desc   ：Excel给单元格新增颜色
=================================================='''
from openpyxl import Workbook

from openpyxl.styles import PatternFill
excel_address = "./test2.xlsx"

wb=Workbook()
wb.create_sheet("Mysheet1", 0)
sht = wb["Mysheet1"]
sht['A1']="Fail"
fille = PatternFill('solid',fgColor='FF0000')
sht['A1'].fill=fille
sht['A2']="Pass"
fille1 = PatternFill('solid',fgColor='238E23')
sht['A2'].fill=fille1
wb.save(excel_address)

