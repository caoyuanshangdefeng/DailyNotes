# -*- coding: UTF-8 -*-
'''=================================================
@Project -> File   ：python_openpyxl -> pyExcel
@IDE    ：PyCharm
@Author ：zhangzhen
@Date   ：2020/5/15 14:58
@Desc   ：
=================================================='''
"""
openpyxl_style_demos.py
（openpyxl二：单元格样式，命名样式；自适应列宽设置）
使用：
#命名样式 及 其应用
#####
#样式_正文样式
#样式_标题行样式
#样式_标题列样式
#自动设置列宽
#####
#使用：for循环遍历得出每列长度后形成字典数据来自动设置每列列宽。
"""

#################### 使用样式styles（单元格样式 和 命名样式NameeStyle/样式模板）:

# 单元格样式：默认格式设置 以及 单元格样式应用:
# openpyxl.style.Font(),         Cell.font,
# openpyxl.style.PatternFill(),  Cell.fill,
# openpyxl.style.Border(),       Cell.border,
# openpyxl.style.Side(),
# openpyxl.style.Alignment(),    Cell.alignment,
# openpyxl.style.Protection();


# 命名样式NameeStyle/样式模板:
# openpyxl.styles.NamedStyle(), Workbook.add_named_style(), Cell.style;
# 注意：
#    命名样式在首次分配给单元时也将自动注册 如ws['A1'].style = NamedStyle(name='namedstyle')
#    注册后，仅使用名称分配样式 如ws['A2'].style=namedstyle


import openpyxl
from openpyxl import Workbook  # 新建工作簿类使用
from openpyxl import load_workbook  # 加载工作簿使用
from openpyxl.worksheet.worksheet import Worksheet  # 工作表单类

from openpyxl.comments import Comment  # 添加注释内容使用
from openpyxl.drawing.image import Image  # 插入图片使用
from openpyxl.worksheet.table import (Table,  # 插入表单使用
                                      TableStyleInfo  # 表单样式信息类
                                      )

from openpyxl.styles import (NamedStyle,  # 使用样式：创建命名样式
                             Font,  # 使用样式：字体，用于设置字体大小、颜色、下划线等
                             PatternFill,  # 使用样式：图样填充
                             Border,  # 使用样式：边框设置
                             Side,  # 使用样式：边框类型设置border_style
                             Alignment,  # 使用样式：对齐方式
                             Protection,  # 使用样式：保护选项
                             colors  # 颜色选项
                             )

wb = Workbook()
wb_filename = r'./openpyxl_styles_demos_case1.xlsx'
ws = wb.create_sheet('样式操作', index=0)

sfont = Font(name='微软雅黑',  # 字体名称 如：微软雅黑、宋体等
             size=12,  # 字号
             bold=False,  # 粗体
             italic=False,  # 斜体
             vertAlign=None,  # 纵向对齐
             underline='none',  # 下划线（‘doubleAccounting’, ‘single’, ‘double’, ‘singleAccounting’）
             strike=False,  # 删除线
             color='FF000000'  # 字体颜色
             )

fill = PatternFill(fill_type='solid',  # 指定填充的类型，支持的有：'solid'等。
                   start_color=colors.YELLOW,  # 指定天聪的开始颜色
                   end_color=colors.YELLOW  # 指定填充的结束颜色
                   )

border = Border(left=Side(border_style='thin', color='FF000000'),  # 左边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                right=Side(border_style='thin', color='FF000000'),  # 右边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                top=Side(border_style='thin', color='FF000000'),  # 上边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                bottom=Side(border_style='thin', color='FF000000'),  # 下边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                diagonal=Side(border_style=None, color='FF000000'),  # 对角线边框设置，Side类定义 边类型/颜色。'thin'/'thick'
                diagonal_direction=0,  # 对角线方向
                outline=Side(border_style=None, color='FF000000'),  # 外边框线设置，Side类定义 变类型/颜色。'thin'/'thick'
                vertical=Side(border_style=None, color='FF000000'),  # 垂直线设置，Side类定义 变类型/颜色。'thin'/'thick'
                horizontal=Side(border_style=None, color='FF000000'),  # 水平线设置，Side类定义 变类型/颜色。'thin'/'thick'
                diagonalDown=False,
                start=None,
                end=None
                )

alignment = Alignment(horizontal='left',
                      # 水平对齐('centerContinuous', 'general', 'distributed','left', 'fill', 'center', 'justify', 'right')
                      vertical='center',  # 垂直对齐（'distributed', 'top', 'center', 'justify', 'bottom'）
                      text_rotation=0,  # 文字旋转
                      wrap_text=True,  # 自动换行
                      shrink_to_fit=False,  # 缩小字体填充
                      mergeCell=None,  # 合并单元格
                      indent=0  # 缩进
                      )

# 命名样式 及 其应用
#####
# 样式_正文样式
style_zhengwen = NamedStyle(name='style_zhengwen',
                            font=Font(),
                            fill=PatternFill(),
                            border=border,  # 所有边框
                            alignment=alignment,  # 水平居中，垂直左对齐，自动换行
                            )

# 样式_标题行样式
style_titleRow = NamedStyle(name='style_titleRow',
                            font=Font(b=True),  # 粗体
                            fill=PatternFill(fill_type='solid',  # 指定填充的类型，支持的有：'solid'等。
                                             start_color='CCCCCC',  # 指定天聪的开始颜色
                                             end_color='CCCCCC'  # 指定填充的结束颜色
                                             ),
                            border=border,  # 所有边框
                            alignment=Alignment(horizontal='center',  # 水平居中
                                                vertical='center',  # 垂直居中
                                                wrap_text=True,  # 自动换行
                                                )
                            )

# 样式_标题列样式
style_titleCol = NamedStyle(name='style_titleCol',
                            font=Font(b=True),  # 粗体
                            fill=PatternFill(fill_type='solid',  # 指定填充的类型，支持的有：'solid'等。
                                             start_color='CCCCCC',  # 指定天聪的开始颜色
                                             end_color='CCCCCC'  # 指定填充的结束颜色
                                             ),
                            border=border,  # 所有边框
                            alignment=Alignment(horizontal='center',  # 水平居中
                                                vertical='center',  # 垂直居中
                                                wrap_text=True,  # 自动换行
                                                )
                            )

# 设置标题行样式
ws.append(('A', 'B', 'C'))
for row in ws['A1:C1']:  # 设置标题行样式
    for cell in row:
        cell.style = style_titleRow

ws['B2'] = 'B2单元格' * 20
ws['B2'].alignment = alignment

ws['C2'] = 'C2单元格数据' * 2
ws['C2'].style = style_zhengwen  # 命名样式的应用

ws['D2'] = 'D2这里有换行符\n接下来又是一行\n然后又是一行\n一行又一行\n一行又一行\n一行又一行\n一行又一行'
ws['D2'].alignment = alignment

ws.merge_cells('F2:J5')  # 合并单元格
ws['F2'] = '合并信息'
ws['F2'].font = sfont  # 单元格样式的应用

# 设置合并的单元格的所有边框
for row in ws['F2:J5']:
    for cell in row:
        cell.style = style_zhengwen

# 设置指定范围单元格的所有边框
for row in ws['A10':'D15']:
    for cell in row:
        cell.border = border

# 手动设置列宽
# ws.column_dimensions['B'].width=20    #设置列宽

# 自动设置列宽
#####
# 使用：for循环遍历得出每列长度后形成字典数据来自动设置每列列宽。
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            """
            首先获取每个单元格中的长度；如果有换行则按单行的长度计算，先分割再计算；
            长度计算中：len('中文')>>>2, len('中文'.encode('utf-8'))>>>6，通过运算，将中文的字节数定义为2；
            字典存储每列的宽度：将cell每列中 列名作为键名，cell长度计算的最大长度作为键值。
            """
            len_cell = max(
                [(len(line.encode('utf-8')) - len(line)) / 2 + len(line) for line in str(cell.value).split('\n')])
            # dims[chr(64+cell.column)] = max((dims.get(chr(64+cell.column), 0), len(str(cell.value))))
            dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len_cell)
for col, value in dims.items():
    """最后通过遍历存储每列的宽度的字典，来设置相关列的宽度"""
    ws.column_dimensions[col].width = value + 2 if value + 2 <= 50 else 50

wb.save(wb_filename)


