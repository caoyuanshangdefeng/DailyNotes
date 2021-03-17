# -*- coding: UTF-8 -*-
'''=================================================
@Project -> File   ：python_openxl -> pyEXCEL_create
@IDE    ：PyCharm
@Author ：zhangzhen
@Date   ：2020/4/20 19:07
@Desc   ：
=================================================='''

from openpyxl import Workbook
#实例化
wb=Workbook()
#激活worksheet
file_path="./testdemo.xlsx"
#方式一：插入到最后(default)
# ws1 = wb.create_sheet("Mysheet1")
# 方式二：插入到最开始的位置
#Mysheet1,Mysheet2作为测试类型名
wb.create_sheet("Mysheet1", 0)
ws2 = wb.create_sheet("Mysheet2", 1)
wb.save(file_path)


