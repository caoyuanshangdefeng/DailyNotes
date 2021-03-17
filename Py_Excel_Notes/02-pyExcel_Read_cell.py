# -*- coding: UTF-8 -*-
'''=================================================
@Project -> File   ：python_openpyxl -> excel_Read_cell
@IDE    ：PyCharm
@Author ：zhangzhen
@Date   ：2020/4/20 20:04
@Desc   ：
=================================================='''
import os


def file_name(file_dir):
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            print(file)  # 当前路径下所有非目录子文件


# print(file_name(r'E:/Daily_use/python_openpyxl/'))
# 打开已有
from openpyxl import load_workbook
wb = load_workbook('./ATtest.xlsx')
print(wb.sheetnames)  # 返回一个列表,sheet名为元素
for sheet in wb:
    print(sheet.title)#遍历打印出所有的sheet名
# 选择sheet名
ws2 = wb['AT_SMS_SendOrder']  # 常用
print('wb2:', ws2)

# 单一单元格访问，获取ＡＴ指令
num = 2
while (1):
    CellInner=ws2['E{}'.format(num)].value
    if CellInner:
        print(CellInner)
    else:
        break
    num=num+1

